# A simple Markov chain model for the weather in Python
import numpy as np
import random as rm
import time
from openpyxl import load_workbook
from regex import P, R

# path is the data of LA from 2017-2022
path = "Excels\\LAweather.xlsx"
# path1 has the data of LA weather from (2022 January 1 - today)
path1 ="Excels\\LApredict.xlsx"
# path2 is the future predicted data
path2 = ("Excels\\LAfuture.xlsx")

wb = load_workbook(path)
wb1 = load_workbook(path1)
wb2 = load_workbook(path2)

sheet = wb.active
sheet1 = wb1.active
sheet2 = wb2.active

#sArray - number of Sunny days
#ssArray - number of Sunny days after Sunny
#scArray - number of Cloudy days after Sunny
#srArray - number of Rainy days after Sunny
# we used this same logic for CC,CS,CR,RR,RS,RC

sArray=[]
ssArray = []
scArray = []
srArray = []

cArray=[]
ccArray = []
csArray = []
crArray = []

rArray=[]
rrArray = []
rsArray = []
rcArray = []


def get():

    for row in range(2, sheet.max_row + 1):
        #calculating overall number of each Sunny, Cloudy, Rainy days
        if sheet[row][1].value == "Sunny":
            sArray.append(row)
        elif sheet[row][1].value == "Cloudy":
            cArray.append(row)
        elif sheet[row][1].value == "Rainy":
            rArray.append(row)
        #Calculating Sunny, Cloudy, Rainy after Sunny days, 
                    #Cloudy, Sunny, Rainy after Cloudy days
                    #Rainy, Sunny, Cloudy after Rainy days.
        if sheet[row][1].value == "Sunny" and sheet[row + 1][1].value == "Sunny":
            ssArray.append(row)
        elif sheet[row][1].value == "Sunny"  and sheet[row + 1][1].value == "Cloudy":
            scArray.append(row)
        elif sheet[row][1].value == "Sunny"  and sheet[row + 1][1].value == "Rainy":
            srArray.append(row)
        elif sheet[row][1].value == "Cloudy" and sheet[row + 1][1].value == "Cloudy":
            ccArray.append(row)
        elif sheet[row][1].value == "Cloudy" and sheet[row + 1][1].value == "Sunny":
            csArray.append(row)
        elif sheet[row][1].value == "Cloudy" and sheet[row + 1][1].value == "Rainy":
            crArray.append(row)
        elif sheet[row][1].value == "Rainy" and sheet[row + 1][1].value == "Rainy":
            rrArray.append(row)
        elif sheet[row][1].value == "Rainy" and sheet[row + 1][1].value == "Sunny":
            rsArray.append(row)
        elif sheet[row][1].value == "Rainy" and sheet[row + 1][1].value == "Cloudy":
            rcArray.append(row)
    #returning len(array) of each Array to see the integer number of elements.
    return len(sArray), len(cArray), len(rArray), len(ssArray), len(scArray), len(srArray), len(ccArray), len(csArray), len(crArray), len(rrArray), len(rsArray), len(rcArray)

lenOfS, lenOfC, lenOfR, lenOfSS, lenOfSC, lenOfSR, lenOfCC, lenOfCS, lenOfCR, lenOfRR, lenOfRS, lenOfRC = get()

print("Sunny days:", lenOfS,'\n' "Cloudy days:", lenOfC, '\n' "Rainy days:", lenOfR)
print('________________________________________________')
print("Sunny-Sunny:", lenOfSS,'\n' "Sunny-Cloudy:", lenOfSC,'\n' "Sunny-Rainy:", lenOfSR)
print('________________________________________________')
print("Cloudy-Cloudy:", lenOfCC,'\n' "Cloudy-Sunny:", lenOfCS,'\n' "Cloudy-Rainy:", lenOfCR)
print('________________________________________________')
print("Rainy-Rainy:", lenOfRR,'\n' "Rainy-Sunny:", lenOfRS,'\n' "Rainy-Cloudy:", lenOfRC)
print('________________________________________________')

# to get the probability of each of them, we divided number of "Sunny days After Sunny" to number of "Sunny" days
# same logic for all
probOfSS = round(lenOfSS/lenOfS,2)
probOfSC = round(lenOfSC/lenOfS,2)
probOfSR = round(lenOfSR/lenOfS,2)

probOfCC = round(lenOfCC/lenOfC,2)
probOfCS = round(lenOfCS/lenOfC,2)
probOfCR = round(lenOfCR/lenOfC,2)

probOfRR = round(lenOfRR/lenOfR,2)
probOfRS = round(lenOfRS/lenOfR,2)
probOfRC = round(lenOfRC/lenOfR,2)

print("Probability of Sunny-Sunny:", probOfSS,'\n' "Probability of Sunny-Cloudy:", probOfSC,'\n' "Probability of Sunny-Rainy:", probOfSR)
print("Probability of Cloudy-Cloudy:", probOfCC,'\n' "Probability of Cloudy-Sunny:", probOfCS,'\n' "Probability of Cloudy-Rainy:", probOfCR)
print("Probability of Rainy-Rainy:", probOfRR,'\n' "Probability of Rainy-Sunny:", probOfRS,'\n' "Probability of Rainy-Cloudy:", probOfRC)


# Let's define the statespace
states = ["Sunny","Cloudy", "Rainy"]

# Possible sequences of events
transitionName = [["SS","SC", "SR"],["CS","CC", "CR"], ["RS","RC", "RR"]]


# Probabilities matrix (transition matrix)
transitionMatrix = [[0.75,0.19,0.06],[0.26,0.6,0.14],[0.3,0.25,0.45]]


# Check that probabilities add to 1. If not, raise ValueError
if sum(transitionMatrix[0])+sum(transitionMatrix[1])+ sum(transitionMatrix[2]) != 3:
    print("Error!!!! Probabilities MUST ADD TO 1. Check transition matrix!!")
    raise ValueError("Probabilities MUST ADD TO 1")


# A functions which implements the Markov model to forecast the weather
def weatherForecast(days):
    # We decided to start from "Cloudy", because the last days data was "Cloudy"
    weatherToday = "Cloudy"
    print("Starting weather: ",weatherToday)
    sheet1[2][2].value= weatherToday
    for row in range (3, days+2):
        if weatherToday == "Sunny":
            #numpy.random.choice(a, size=None, replace=True, p=None)
            change = np.random.choice(transitionName[0],replace=True,p=transitionMatrix[0])
            if change == "SS":
                pass
            elif change== "SC":
                weatherToday = "Cloudy"
            else: 
                weatherToday= "Rainy"

        elif weatherToday == "Cloudy":
            change = np.random.choice(transitionName[1],replace=True,p=transitionMatrix[1])
            if change == "CC":
                pass
            elif change == "CS":
                weatherToday = "Sunny"
            else:
                weatherToday= "Rainy"
        elif weatherToday == "Rainy":
            change = np.random.choice(transitionName[2], replace=True, p=transitionMatrix[2])
            if change =="RR":
                pass
            elif change == "RS":
                weatherToday= "Sunny"
            else:
                weatherToday = "Cloudy"
        print(weatherToday)
        sheet1[row][2].value = weatherToday
        sheet2[row][1].value = weatherToday
        

# # We forecast the weather for 143 day, to compare historical data from 2022 January 1 to today, to the data we predicted
# weatherForecast(143)

## We forcast weather for the upcoming 5 years(1828 days)
weatherForecast(1828)

# wb1.save(path1)
# wb2.save(path2)
