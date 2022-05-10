from openpyxl import load_workbook
path1 ="C:\\Users\\ktiki\\Desktop\\LApredict.xlsx"
wb= load_workbook(path1)
sheet = wb.active
matched = 0
for row in range (2, sheet.max_row+1):
    if sheet[row][1].value ==  sheet[row][2].value and sheet[row][1].value != None:
        matched+=1

print(matched)
#we have historical data from (January 1 - today) - overall 143 results
#out of 143 we have 123 equal to each other
